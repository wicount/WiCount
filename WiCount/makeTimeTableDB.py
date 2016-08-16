#Imports correct libraries on the server
#import sys
#sys.path.insert(0, "/home/student/anaconda3/lib/python3.4/site-packages")

import db
import sqlite3 as lite
from sqlite3 import OperationalError
import glob, os
import openpyxl
import numpy as np
import wicount
import datetime
from dateutil.parser import parse

def GetRoomNo(room):
    ''' format the room number into the correct format'''
    room_no = room[:1] + "-"+ room[1:2] + room [3:]
    return room_no

def GetDay(i):
    ''' return the day of the week'''
    return {
        1: "Mon",
        3: "Tue",
        5: "Wed",
        7: "Thu",
        9: "Fri",
    }.get(i, 1)

def GetWeekNo(week):
    ''' get the week number'''
    date = week.split("-")[0]
    date = parse(date)
    week1 = date.isocalendar()
    week_nos = [str(week1[1]) + "/" + str(week1[0]), str(week1[1] +1) + "/" + str(week1[0])]
    return week_nos
   
def AddDetailsToTimeTable(line, weekNo,room_id):
    ''' add the details to the database along with the week number '''
    sqlvalues = []
    module_codes = []
    start_time = wicount.GetTime(line[0])
    #build sql string
    for i in range(1,len(line),2):
        db_values = [room_id, GetDay(i), start_time, weekNo, line[i], line[i+1]]
        if line[i] and line[i+1] and line[i+1] != "N/A":
            module = [line[i], line[i+1]]
            module_codes.append(module)
        sqlvalues.append(db_values)
    try:
        c.executemany('INSERT OR REPLACE INTO timetable VALUES (?,?,?,?,?,?)', sqlvalues)
    except OperationalError:
        print ("Command skipped: ", sqlvalues)
    con.commit()
    return module_codes

def UpdateModuleTable(module_list, week_no):    
    ''' update the modules table'''
    # get unique values from the module list
    #http://stackoverflow.com/questions/13464152/typeerror-unhashable-type-list-when-using-built-in-set-function
    module_list = sorted(set(map(tuple, module_list)))
    sqlvalues = []
    
    for x in range(0, len(module_list)):
        if x != len(module_list)-1 and module_list[x][0] == module_list[x+1][0]:
            continue
        else:
            values = [module_list[x][0], week_no, module_list[x][1]]
            sqlvalues.append(values)

    try:            
        c.executemany('INSERT OR REPLACE INTO modules VALUES (?,?,?)', sqlvalues)
    except OperationalError:
        print ("Command skipped: ", sqlvalues)
    con.commit()
     
con = db.get_connection()
c=con.cursor()

def main():      
    # Got help from http://stackoverflow.com/questions/3964681/find-all-files-in-directory-with-extension-txt-in-python
    os.chdir("timetable")
    
    #set up hard codeing this will need to be passed in.
    campus = "Belfield"
    building = "Computer Science"
    
    #Extract Timetable data from Excel file
    for file in glob.glob("*.xlsx"):
        wb = openpyxl.load_workbook(file)
        sheetnames = wb.get_sheet_names()
        sqlvalues = []
        module_list = []
        for sheet in sheetnames:
            if sheet != "All":
                sheetData = wb.get_sheet_by_name(sheet)
                
                #get the room capacity from the spread sheet
                capacity = sheetData.cell('C1','C1').value
                capacity = capacity.split()[2]
                
                # get the week number for the timetable
                week1 = sheetData.cell('B1','B1').value
                week_nos = GetWeekNo(week1)
                week1 = week_nos[0]
                week2 = week_nos[1]
                
                room_details = [campus, building, GetRoomNo(sheet),int(capacity)] 
                room_id = wicount.GetRoomID(room_details)
                
                table = np.array([[cell.value for cell in col] for col in sheetData['A3':'K11']])
                for line in table:
                    module_list.extend(AddDetailsToTimeTable(line, week1,room_id))
                
                table = np.array([[cell.value for cell in col] for col in sheetData['M3':'W11']])
                for line in table:
                    module_list.extend(AddDetailsToTimeTable(line, week2, room_id))
            #end if
        #end for
        os.remove(file)
                    
        UpdateModuleTable(module_list, week1)
    #end for loop 
    # con.close() 
    os.chdir("../") #Return to original directory
    print("finished MakeTimeTable") 
